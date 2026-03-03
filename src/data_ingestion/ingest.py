"""
Azure AI Search - Data Ingestion Script
Indexes Excel (.xlsx), Word (.docx) and PDF (.pdf) files from a sample folder
into an Azure AI Search index.
"""

import hashlib
import logging
import os
import sys
from pathlib import Path

import fitz  # PyMuPDF
from azure.identity import DefaultAzureCredential
from azure.search.documents import SearchClient
from azure.search.documents.indexes import SearchIndexClient
from azure.search.documents.indexes.models import (
    SearchableField,
    SearchField,
    SearchFieldDataType,
    SearchIndex,
    SimpleField,
)
from docx import Document as DocxDocument
from dotenv import load_dotenv
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
load_dotenv()

SEARCH_ENDPOINT = os.environ["AZURE_SEARCH_ENDPOINT"]
SEARCH_INDEX_NAME = os.getenv("AZURE_SEARCH_INDEX_NAME", "documents-index")
# If set, key-based auth is used; otherwise DefaultAzureCredential (recommended)
SEARCH_API_KEY = os.getenv("AZURE_SEARCH_API_KEY")

SAMPLE_FOLDER = Path(__file__).parent / "sample"

SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".xlsx"}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Credential helper
# ---------------------------------------------------------------------------
def _get_credential():
    """Return an API-key credential dict or a DefaultAzureCredential."""
    if SEARCH_API_KEY:
        from azure.core.credentials import AzureKeyCredential

        return AzureKeyCredential(SEARCH_API_KEY)
    return DefaultAzureCredential()


# ---------------------------------------------------------------------------
# Text extraction helpers
# ---------------------------------------------------------------------------
def extract_text_from_pdf(file_path: Path) -> str:
    """Extract text from a PDF file using PyMuPDF."""
    text_parts: list[str] = []
    with fitz.open(str(file_path)) as doc:
        for page in doc:
            text_parts.append(page.get_text())
    return "\n".join(text_parts)


def extract_text_from_docx(file_path: Path) -> str:
    """Extract text from a Word .docx file."""
    doc = DocxDocument(str(file_path))
    return "\n".join(paragraph.text for paragraph in doc.paragraphs if paragraph.text)


def extract_text_from_xlsx(file_path: Path) -> str:
    """Extract text from an Excel .xlsx file (all sheets)."""
    wb = load_workbook(str(file_path), read_only=True, data_only=True)
    text_parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        text_parts.append(f"--- Sheet: {sheet_name} ---")
        for row in ws.iter_rows(values_only=True):
            row_text = "\t".join(
                str(cell) if cell is not None else "" for cell in row)
            if row_text.strip():
                text_parts.append(row_text)
    wb.close()
    return "\n".join(text_parts)


EXTRACTORS = {
    ".pdf": extract_text_from_pdf,
    ".docx": extract_text_from_docx,
    ".xlsx": extract_text_from_xlsx,
}


# ---------------------------------------------------------------------------
# Index management
# ---------------------------------------------------------------------------
def create_or_update_index(index_client: SearchIndexClient) -> None:
    """Create (or update) the search index with the expected schema."""
    fields = [
        SimpleField(
            name="id",
            type=SearchFieldDataType.String,
            key=True,
            filterable=True,
        ),
        SearchableField(
            name="content",
            type=SearchFieldDataType.String,
            analyzer_name="fr.microsoft",
        ),
        SimpleField(
            name="file_name",
            type=SearchFieldDataType.String,
            filterable=True,
            sortable=True,
        ),
        SimpleField(
            name="file_type",
            type=SearchFieldDataType.String,
            filterable=True,
            facetable=True,
        ),
        SearchableField(
            name="title",
            type=SearchFieldDataType.String,
            analyzer_name="fr.microsoft",
        ),
    ]

    index = SearchIndex(name=SEARCH_INDEX_NAME, fields=fields)
    result = index_client.create_or_update_index(index)
    logger.info("Index '%s' created / updated successfully.", result.name)


# ---------------------------------------------------------------------------
# Document ingestion
# ---------------------------------------------------------------------------
def _generate_document_id(file_path: Path) -> str:
    """Generate a stable, URL-safe document id from the file path."""
    return hashlib.sha256(file_path.name.encode("utf-8")).hexdigest()[:32]


def collect_documents(folder: Path) -> list[dict]:
    """Walk the sample folder and extract text from supported files."""
    documents: list[dict] = []
    if not folder.exists():
        logger.error("Sample folder not found: %s", folder)
        sys.exit(1)

    for file_path in sorted(folder.iterdir()):
        ext = file_path.suffix.lower()
        if ext not in SUPPORTED_EXTENSIONS:
            logger.warning("Skipping unsupported file: %s", file_path.name)
            continue

        logger.info("Processing %s …", file_path.name)
        extractor = EXTRACTORS[ext]
        try:
            content = extractor(file_path)
        except Exception:
            logger.exception("Failed to extract text from %s", file_path.name)
            continue

        doc = {
            "id": _generate_document_id(file_path),
            "content": content,
            "file_name": file_path.name,
            "file_type": ext.lstrip("."),
            "title": file_path.stem,
        }
        documents.append(doc)

    return documents


def upload_documents(search_client: SearchClient, documents: list[dict]) -> None:
    """Upload documents to the Azure AI Search index in batches."""
    if not documents:
        logger.warning("No documents to upload.")
        return

    BATCH_SIZE = 100
    for i in range(0, len(documents), BATCH_SIZE):
        batch = documents[i: i + BATCH_SIZE]
        results = search_client.upload_documents(batch)
        succeeded = sum(1 for r in results if r.succeeded)
        failed = len(batch) - succeeded
        logger.info(
            "Batch %d–%d: %d succeeded, %d failed.",
            i + 1,
            i + len(batch),
            succeeded,
            failed,
        )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    credential = _get_credential()

    # 1. Create / update the index
    index_client = SearchIndexClient(
        endpoint=SEARCH_ENDPOINT, credential=credential
    )
    create_or_update_index(index_client)

    # 2. Collect documents from the sample folder
    documents = collect_documents(SAMPLE_FOLDER)
    logger.info("Found %d document(s) to index.", len(documents))

    # 3. Upload to the index
    search_client = SearchClient(
        endpoint=SEARCH_ENDPOINT,
        index_name=SEARCH_INDEX_NAME,
        credential=credential,
    )
    upload_documents(search_client, documents)

    logger.info("Done – ingestion complete.")


if __name__ == "__main__":
    main()
